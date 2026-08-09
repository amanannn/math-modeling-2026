"""
2023B多波束测线 第二问计算
β: 测线方向与坡面法向水平投影的夹角; L: 测线上距中心点距离(海里)
H(L,β) = H0 - L·tanα·cosβ
tanα' = tanα·|sinβ|
W = H·[1/(tan30°+tanα') + 1/(tan30°-tanα')]
用法: python solve_q2.py [--write]
"""
import argparse
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parent.parent
RESULT = ROOT / 'result2.xlsx'

H0 = 120.0          # 中心点水深 (m)
ALPHA = np.deg2rad(1.5)
NM = 1852.0         # 1海里 = 1852m
BETA = np.array([0, 45, 90, 135, 180, 225, 270, 315])
L_NM = np.array([0, 0.3, 0.6, 0.9, 1.2, 1.5, 1.8, 2.1])

t30 = np.tan(np.pi / 6)
ta = np.tan(ALPHA)


def width(H, ta_p):
    return H * (1 / (t30 + ta_p) + 1 / (t30 - ta_p))


def main(write=False):
    W = np.zeros((len(BETA), len(L_NM)))
    for i, b in enumerate(BETA):
        cb, sb = np.cos(np.deg2rad(b)), np.sin(np.deg2rad(b))
        ta_p = ta * abs(sb)                     # 视坡角
        for j, lnm in enumerate(L_NM):
            H = H0 - lnm * NM * ta * cb         # 沿测线水深
            W[i, j] = width(H, ta_p)

    print(f"{'β\\L(海里)':>9}" + "".join(f"{l:>9.1f}" for l in L_NM))
    for i, b in enumerate(BETA):
        print(f"{b:>9.0f}" + "".join(f"{W[i, j]:>9.2f}" for j in range(len(L_NM))))

    if write:
        from openpyxl import load_workbook
        wb = load_workbook(RESULT)
        ws = wb.active
        for i in range(len(BETA)):
            for j in range(len(L_NM)):
                ws.cell(3 + i, 3 + j).value = round(W[i, j], 2)
        wb.save(RESULT)
        print(f'[OK] 写入 {RESULT}')


if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('--write', action='store_true')
    main(write=p.parse_args().write)
