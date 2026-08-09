"""
2023B多波束测线 第一问计算
D0=70m, α=1.5°, 开角120°(边缘与竖直60°), 测线间距d=200m
W = D·[1/(tan30°+tanα) + 1/(tan30°-tanα)]
η = 1 - d/W_prev
用法: python solve_q1.py [--write]
"""
import argparse
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parent.parent   # 单刷2023B/
RESULT = ROOT / 'result1.xlsx'

D0 = 70.0
ALPHA = np.deg2rad(1.5)
D_LINE = 200.0
X = np.array([-800, -600, -400, -200, 0, 200, 400, 600, 800])


def depth(x):
    return D0 - x * np.tan(ALPHA)          # x正方向为坡上(浅侧)


def width(D):
    t30 = np.tan(np.pi / 6)
    return D * (1 / (t30 + np.tan(ALPHA)) + 1 / (t30 - np.tan(ALPHA)))


def main(write=False):
    Ds = depth(X)
    Ws = width(Ds)
    print(f"{'位置/m':>7} {'水深/m':>9} {'W/m':>9} {'重叠率':>8}")
    for i, (x, d, w) in enumerate(zip(X, Ds, Ws)):
        eta = '——' if i == 0 else (
            f"{1 - D_LINE/Ws[i]:.2%}" if 1 - D_LINE/Ws[i] >= 0 else '无重叠')
        print(f"{x:>7} {d:>9.2f} {w:>9.2f} {eta:>8}")

    if write:
        from openpyxl import load_workbook
        wb = load_workbook(RESULT)
        ws = wb.active
        for i in range(len(X)):
            ws.cell(2, i + 2).value = round(Ds[i], 2)
            ws.cell(3, i + 2).value = round(Ws[i], 2)
            if i > 0:
                e = 1 - D_LINE / Ws[i]
                ws.cell(4, i + 2).value = '无重叠' if e < 0 else f'{e:.2%}'
        wb.save(RESULT)
        print(f'[OK] 写入 {RESULT}')


if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('--write', action='store_true')
    main(write=p.parse_args().write)
